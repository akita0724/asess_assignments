from csv import reader as csv_reader
from distutils.file_util import copy_file
from genericpath import isfile
from re import findall
from subprocess import Popen, PIPE, TimeoutExpired
from os import mkdir, path, listdir, cpu_count
from time import time
from multiprocessing import Pool
from shutil import which as shutil_which
try:
    from openpyxl import Workbook
except:
    raise EnvironmentError("Package 'openpyxl' not found.")

tqdmUsable = False
try:
    from tqdm import tqdm
    tqdmUsable = True
except:
    print("NOTICE")
    print("Package 'tqdm' not installed. In order to show progress, please install it.")
    print()
    pass


class JudgeProgram:

    report = []
    reportDetail = []

    def __init__(self) -> None:

        if shutil_which("python") is not None:
            self.pythonPath = shutil_which("python")
        elif shutil_which("python3") is not None:
            self.pythonPath = shutil_which("python3")
        else:
            raise EnvironmentError("Python not found.")

        while True:
            self.AssignfolderName: str = input("Assignments Directory --> ")

            if path.isdir(self.AssignfolderName) == False:
                if path.isfile(self.AssignfolderName) == False:
                    print("Given string is not a directory.")
                else:
                    print("Directory not found.")
            else:
                break

        while True:
            self.resultDir: str = input("Result Directory --> ")

            if path.isfile(self.resultDir) == True:
                print("Given string is not a directory.")
            elif path.isdir(self.resultDir) == False:
                mkdir(self.resultDir)
                break
            else:
                print("Directory already exists.")

        while True:
            self.stdInFileName: str = input("Stdin File Name --> ")

            if path.isfile(self.stdInFileName) == False:
                print("Stdin file not found.")
            elif path.splitext(self.stdInFileName)[1].lower() != ".csv":
                print("Stdin file must be csv.")
            else:
                break

        while True:
            self.answerFileName: str = input("Answer File Name --> ")

            if path.isfile(self.answerFileName) == False:
                print("Answer file not found.")
            elif path.splitext(self.answerFileName)[1].lower() != ".csv":
                print("Answer file must be csv.")
            else:
                break

        self.fileNames, self.file_unceked = self.getFiles(
            self.AssignfolderName)

        self.fileAmount: int = len(self.fileNames)

        if self.fileAmount == 0:
            raise FileNotFoundError("Files not found.")

        while True:
            value = input("Timeout (s) --> ")
            if value == "":
                self.timeout = 3
                break
            else:
                try:
                    self.timeout = int(value)
                    break
                except:
                    continue

        while True:
            value = input("Max Process --> ")
            cpu_: int = cpu_count() if cpu_count() is not None else 1  # type: ignore
            if value == "":
                self.max_process = cpu_
                break
            else:
                try:
                    self.max_process = int(value)
                    if self.max_process < 1:
                        print("Max process must be 1 or more.")
                    if self.max_process > cpu_:
                        print(
                            "Max process must be less than or equal to the number of CPUs.")
                    break
                except:
                    continue

        mkdir(path.join(self.resultDir, "AC"))
        mkdir(path.join(self.resultDir, "WA"))
        mkdir(path.join(self.resultDir, "RE"))
        mkdir(path.join(self.resultDir, "TLE"))
        mkdir(path.join(self.resultDir, "unchecked"))

    def getFiles(self, folderName: str) -> tuple[list[str], list[str]]:
        fileNames: list[str] = listdir(folderName)

        file_checked: list[str] = []
        file_unchecked: list[str] = []
        for fileName in fileNames:
            if path.isfile(path.join(folderName, fileName)) and path.splitext(fileName)[1].lower() == ".py":
                file_checked.append(fileName)
            else:
                file_unchecked.append(fileName)

        return file_checked, file_unchecked

    def getStdIn(self, fileName: str) -> None:
        with open(fileName, encoding="utf-8") as f:
            reader = csv_reader(f)
            l = ['\n'.join(row) for row in reader]
        self.stdIns = l
        self.stdInAmount = len(l)

    def getAnswer(self, fileName: str) -> None:
        with open(fileName, encoding="utf-8") as f:
            reader = csv_reader(f)
            l = [row for row in reader]
        self.answers = l
        self.answerAmount = len(l)

    def judge(self, testId: int, arg) -> list[str | int]:

        answer, stdin, timeout, filePath, fileIndex = arg

        start = time()

        process = Popen([self.pythonPath, filePath], stdin=PIPE,  # type: ignore
                        stdout=PIPE, stderr=PIPE, text=True)  # type: ignore

        end = time()

        judgeId = "-".join([str(fileIndex), str(testId)])

        try:
            stdout, stderr = process.communicate(
                stdin, timeout=timeout)

        except TimeoutExpired:
            process.kill()
            # TLE
            return [judgeId, 4, "", timeout * 1000]

        stdout_splitted = stdout.split("\n")

        returncode = process.returncode

        exe_time = int((end - start) * 1000)

        if returncode == 0:
            # AC
            if stdout_splitted == answer:
                return [judgeId, 0, "", exe_time]
            # WA
            else:
                return [judgeId, 1, "", exe_time]
        # RE
        else:
            return [judgeId, 2, stderr, exe_time]

    def processL(self, fileName: str, fileTo: str, ave_exe: float, status: str) -> None:

        copy_file(path.join(self.AssignfolderName, fileName),
                  path.join(self.resultDir, fileTo, fileName))

        fileSize = path.getsize(path.join(self.AssignfolderName, fileName))

        self.report.append([self.getStudentNumber(
            fileName), status, ave_exe, fileSize])

    def getStudentNumber(self, fileName: str) -> str:
        def verify_number(number: str) -> tuple[bool, str]:
            match = findall("[1-3][0-9][0-4][0-9]", number)
            match.append(number)
            return len(match) != 0, match[0]
        print(fileName)
        candicate = []
        candicate.extend(fileName.split("_"))
        candicate.extend(fileName.split("-"))

        for c in candicate:
            print(c)
            res, c = verify_number(c)
            if res:
                return c

        return f"U-{fileName}"

    def judgeFile(self, name: str, fileIndex: int) -> None:
        results = []

        a = tuple(enumerate(zip(self.answers, self.stdIns,
                                [self.timeout] *
                                self.stdInAmount,
                                [path.join(self.AssignfolderName, name)
                                 ] * self.stdInAmount,
                                [fileIndex] * self.stdInAmount)))

        with Pool(self.max_process) as pool:
            results = pool.starmap(
                self.judge, a)

        status = [0, 0, 0, 0]
        ave_sum = 0
        for result in results:
            status[result[1]] += 1  # type: ignore
            ave_sum += result[3]  # type: ignore

            self.reportDetail.append(result)

        ave_exe = ave_sum / self.stdInAmount

        if status[0] == self.stdInAmount:
            judgeStatus = "AC"
        elif status[2] > 0:
            judgeStatus = "RE"
        elif status[1] > 0:
            judgeStatus = "WA"
        else:
            judgeStatus = "TLE"

        self.processL(name, judgeStatus, ave_exe, judgeStatus)

    def moveUncheck(self) -> None:
        for fileName in self.file_unceked:
            copy_file(path.join(self.AssignfolderName, fileName),
                      path.join(self.resultDir, "unchecked", fileName))

    def writeReport(self) -> None:
        wb = Workbook()

        wb.create_sheet("判定詳細", 0)
        wb_detail = wb["判定詳細"]
        wb_detail.cell(1, 1, "判定ID")
        wb_detail.cell(1, 2, "判定結果")
        wb_detail.cell(1, 3, "エラーメッセージ")
        wb_detail.cell(1, 4, "実行時間(ms)")
        for r, row in enumerate(self.reportDetail):
            for c, cell in enumerate(row):
                wb_detail.cell(r + 2, c + 1, cell)

        wb_detail.cell(1, 6, "判定結果の凡例")
        wb_detail.cell(2, 6, "0")
        wb_detail.cell(2, 7, "AC : 正解")
        wb_detail.cell(3, 6, "1")
        wb_detail.cell(3, 7, "WA : 不正解")
        wb_detail.cell(4, 6, "2")
        wb_detail.cell(4, 7, "RE : 実行時エラー")
        wb_detail.cell(5, 6, "3")
        wb_detail.cell(5, 7, "TLE : 実行時間制限超過")

        wb.create_sheet("判定結果一覧", 1)
        wb_result = wb["判定結果一覧"]

        wb_result.cell(1, 1, "学籍番号")
        wb_result.cell(1, 2, "判定結果")
        wb_result.cell(1, 3, "平均実行時間(ms)")
        wb_result.cell(1, 4, "ファイルサイズ(byte)")
        for r, row in enumerate(self.report):
            for c, cell in enumerate(row):
                wb_result.cell(r + 2, c + 1, cell)

        wb.save(path.join(self.resultDir, "result.xlsx"))
        del wb

    def main(self) -> None:
        self.getStdIn(self.stdInFileName)
        self.getAnswer(self.answerFileName)

        if self.stdInAmount != self.answerAmount:
            raise ValueError(
                "The number of stdin and the number of answers do not match.")

        print("Judge started.")

        if tqdmUsable:
            for i, fileName in enumerate(tqdm(self.fileNames)):
                self.judgeFile(fileName, i + 1)
        else:
            for i, fileName in enumerate(self.fileNames):
                self.judgeFile(fileName, i + 1)

        self.moveUncheck()

        print("Judge finished. Creating report...")

        self.writeReport()

        print("Report created.")


if __name__ == "__main__":
    judgeProgram = JudgeProgram()
    judgeProgram.main()
